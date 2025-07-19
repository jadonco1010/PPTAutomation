import os
import logging
import time
from flask import Flask, request, render_template_string, send_file
from datetime import datetime
from pathlib import Path
import tempfile
from urllib.parse import unquote # Keep this for parsing if needed, though direct upload simplifies

# Import your custom modules
import config
import excel_processor
import ppt_updater
import utils

# Configure logging for the Flask app
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

app = Flask(__name__)

# Define a temporary directory for file operations within the container
# This will be used for uploaded Excel and generated PowerPoint files
TEMP_DIR = Path(tempfile.gettempdir()) / "app_data"
TEMP_DIR.mkdir(parents=True, exist_ok=True)
logging.info(f"Temporary application data directory: {TEMP_DIR}")

# HTML for the simple upload form
UPLOAD_FORM_HTML = """
<!doctype html>
<title>Excel to PowerPoint Automation</title>
<h1>Upload Excel File</h1>
<form method=post enctype=multipart/form-data action="/upload">
  <input type=file name=excel_file>
  <input type=submit value=Upload>
</form>
{% if message %}
<p>{{ message }}</p>
{% endif %}
"""

@app.route('/')
def index():
    """Serves the main upload form page."""
    return render_template_string(UPLOAD_FORM_HTML)

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handles the uploaded Excel file, processes it, and returns a PowerPoint."""
    start_time = time.time()
    logging.info('Received file upload request.')

    # Check if a file was uploaded
    if 'excel_file' not in request.files:
        logging.error("No file part in the request.")
        return render_template_string(UPLOAD_FORM_HTML, message="No file selected!"), 400

    excel_file = request.files['excel_file']

    # If the user does not select a file, the browser submits an
    # empty file without a filename.
    if excel_file.filename == '':
        logging.error("No selected file.")
        return render_template_string(UPLOAD_FORM_HTML, message="No selected file!"), 400

    if excel_file:
        # Securely save the uploaded file to a temporary location
        # Use a sanitized filename to prevent path traversal issues
        original_filename = excel_file.filename
        # For simplicity, just use the original filename, but in production,
        # consider using werkzeug.utils.secure_filename
        temp_input_excel_path = TEMP_DIR / original_filename
        excel_file.save(temp_input_excel_path)
        logging.info(f"Excel file saved to: {temp_input_excel_path}")

        try:
            # --- Main Logic (adapted from your function_app.py) ---

            # Step 1: Preprocess the Excel file
            # Note: excel_processor.preprocess_excel_xml saves to config.TARGET_EXCEL_FILENAME
            # We need to ensure config.TARGET_EXCEL_FILENAME is also within TEMP_DIR
            config.TARGET_EXCEL_FILENAME = TEMP_DIR / "preprocessed_data.xlsx"
            logging.info(f"Starting preprocessing of {temp_input_excel_path} to {config.TARGET_EXCEL_FILENAME}")
            excel_processor.preprocess_excel_xml(temp_input_excel_path, config.TARGET_EXCEL_FILENAME)
            logging.info("Excel preprocessing completed.")

            # Load the preprocessed Excel file (data_only=True)
            from openpyxl import load_workbook # Import here to avoid potential circular dependency issues
            wb_data_only = load_workbook(config.TARGET_EXCEL_FILENAME, data_only=True, read_only=True)
            sheet_names = wb_data_only.sheetnames
            logging.info(f"Sheets in preprocessed Excel: {sheet_names}")

            current_date = datetime.now()
            fiscal_year, fiscal_quarter_str, fiscal_month_overall_str, fiscal_month_in_quarter_str = utils.get_fiscal_quarter_and_month(current_date.date())
            logging.info(f"Current fiscal period for PPT update: {fiscal_quarter_str} {fiscal_month_overall_str} (Month in Q: {fiscal_month_in_quarter_str})")

            # Dynamic PowerPoint filename generation
            fiscal_year_short = str(fiscal_year)[2:]
            dynamic_filename_part = f"{fiscal_month_in_quarter_str} {fiscal_quarter_str}FY{fiscal_year_short}"
            new_ppt_filename = f"{dynamic_filename_part} P&L Review_Cisco Highly Confidential _WD-1 DRAFT.pptx"
            dynamic_final_output_ppt_path = TEMP_DIR / new_ppt_filename # Ensure output is in TEMP_DIR
            logging.info(f"Dynamic final PowerPoint output path set to: '{dynamic_final_output_ppt_path}'.")

            # Dynamically determine the required sheet names for PPT update
            identified_sheets = utils.find_dynamic_sheets(sheet_names, fiscal_quarter_str, fiscal_month_overall_str, fiscal_month_in_quarter_str)
            logging.info(f"Sheets identified for PPT update: {identified_sheets}")

            # Define the desired order of sheets for table processing
            desired_order_bases = ["Exec View", "Comparisons", "Commit", "Margins Scenarios"]
            ordered_sheets = []

            # Map dynamically identified sheets to their corresponding positions in the desired order
            for desired_base_name in desired_order_bases:
                matched_sheet = next((s for s in identified_sheets if desired_base_name in s), None)
                if matched_sheet:
                    ordered_sheets.append(matched_sheet)
                else:
                    logging.warning(f"Sheet matching base name '{desired_base_name}' not found for PPT update. It will be skipped.")

            logging.info(f"Ordered sheets for PPT update: {ordered_sheets}")

            if len(ordered_sheets) < len(desired_order_bases):
                missing_bases = set(desired_order_bases) - set(s.split(' ', 1)[-1] if ' ' in s else s for s in ordered_sheets)
                logging.critical(f"Missing critical sheets for PPT update: {', '.join(missing_bases)}. Ensure the source file is correct and named as expected.")
                # You might want to return an error here instead of proceeding
                # return render_template_string(UPLOAD_FORM_HTML, message=f"Missing critical sheets: {', '.join(missing_bases)}"), 500


            # Define the base template for table regions (as in your original main.py)
            base_regions = {
                "Exec View": [
                    ("C3", "E13"), ("F3", "F13"),
                    ("C18", "E24"), ("F18", "F24"), ("H18", "H18"), ("H20", "H22"),
                    ("C29", "E36"), ("F29", "F36"), ("H29", "H29"), ("H31", "H33"),
                    ("K3", "K3"), ("K4", "N13"), ("O3", "O3"), ("R4", "R13"),
                    ("S3", "S3"), ("V4", "V13"), ("W3", "W3"), ("W4", "W13"),
                ],
                "Comparisons": [
                    ("K3", "K3"), ("K4", "N13"), ("S3", "S13"), ("T3", "T3"),
                    ("W4", "W13"), ("X3", "X13"), ("AC3", "AC13"), ("AD3", "AD3"),
                    ("AG4", "AG13"), ("AH3", "AH13"), ("AM3", "AM13"), ("AN3", "AN3"), ("AQ4", "AQ13"),
                    ("AR3", "AR13"),
                ],
                "Commit": [
                    ("C3", "C3"), ("C4", "F13"), ("G3", "G3"), ("J4", "J13"),
                    ("K3", "K3"), ("N4", "N13"), ("O3", "O3"), ("R4", "R13"),
                    ("S3", "S3"), ("V4", "V13"),
                ],
                "Margins Scenarios": [
                    ("B15", "B15"), ("B16", "G19"), ("B20", "G20"), ("B25", "B25"),
                    ("B26", "G29"), ("B30", "G30"), ("B32", "B32"), ("B33", "G36"),
                    ("B37", "G37"), ("B39", "B39"), ("B40", "G43"), ("B44", "G44"),
                    ("B46", "B46"), ("B47", "G50"), ("B51", "G51"), ("I39", "I39"),
                    ("I40", "N43"), ("I44", "N44"),
                ],
            }

            table_regions = {}
            for sheet_name in ordered_sheets:
                base_sheet_name = next((k for k in base_regions if k in sheet_name), None)
                if base_sheet_name:
                    table_regions[sheet_name] = base_regions[base_sheet_name]

            logging.info("Loading tables from preprocessed Excel for PowerPoint update...")
            tables = ppt_updater.load_tables_from_excel(str(config.TARGET_EXCEL_FILENAME), table_regions.keys(), table_regions, wb_data_only)
            logging.info("Tables loaded successfully.")

            # Step 3: Update the PowerPoint presentation
            logging.info("Updating PowerPoint presentation...")
            ppt_updater.update_ppt_labels(str(config.PPT_TEMPLATE_FILENAME), str(dynamic_final_output_ppt_path), tables)
            logging.info(f"PowerPoint updated and saved to {dynamic_final_output_ppt_path}")

            end_time = time.time()
            elapsed_time = end_time - start_time
            logging.info(f"Script completed in {elapsed_time:.2f} seconds.")

            # Return the generated PowerPoint file for download
            return send_file(str(dynamic_final_output_ppt_path),
                             mimetype='application/vnd.openxmlformats-officedocument.presentationml.presentation',
                             as_attachment=True,
                             download_name=new_ppt_filename)

        except Exception as e:
            logging.error(f"An error occurred during PowerPoint generation: {e}", exc_info=True)
            return render_template_string(UPLOAD_FORM_HTML, message=f"An error occurred: {e}"), 500
    
    # Fallback if no file was processed (should not be reached with checks above)
    return render_template_string(UPLOAD_FORM_HTML, message="Something went wrong with the file upload."), 500

if __name__ == '__main__':
    # For local testing, CAE will run this with Gunicorn
    app.run(debug=True, host='0.0.0.0', port=5000) # Use 5000 for local Flask dev server