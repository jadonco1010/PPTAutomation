import zipfile
import xml.etree.ElementTree as ET
import logging
from typing import Set, List, Dict, Tuple, Any
from openpyxl import Workbook
from datetime import datetime

from utils import coordinate_to_tuple, find_dynamic_sheets, get_fiscal_quarter_and_month

def _find_sheet_xml_path(z: zipfile.ZipFile, sheet_name: str) -> str:
    """Helper to find the XML path for a given sheet within the Excel zip archive."""
    ns_wb = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    wb_xml = ET.fromstring(z.read("xl/workbook.xml"))
    sh = wb_xml.find(f".//ns:sheet[@name='{sheet_name}']", ns_wb)
    if sh is None:
        raise KeyError(f"Sheet '{sheet_name}' not found in workbook.xml.")
    rid = sh.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]

    ns_rel = {"pr": "http://schemas.openxmlformats.org/package/2006/relationships"}
    rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    entry = rels.find(f".//pr:Relationship[@Id='{rid}']", ns_rel)
    if entry is None:
        raise KeyError(f"Relationship {rid} not in workbook.xml.rels.")
    target = entry.attrib["Target"]
    if not target.startswith("xl/"):
        target = "xl/" + target
    return target

def _get_hidden_rows_from_xml(file_path: str, sheet_name: str) -> Set[int]:
    """Extracts hidden row numbers for a specific sheet by parsing its XML."""
    with zipfile.ZipFile(file_path) as z:
        try:
            xml_path = _find_sheet_xml_path(z, sheet_name)
            tree = ET.fromstring(z.read(xml_path))
            ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            hidden_rows = {
                int(r.attrib["r"])
                for r in tree.findall(".//ns:row", ns)
                if r.get("hidden") == "1"
            }
            return hidden_rows
        except KeyError as e:
            logging.warning(f"Could not find XML path for sheet '{sheet_name}' or related entry: {e}")
            return set()
        except Exception as e:
            logging.error(f"Error getting hidden rows for sheet '{sheet_name}': {e}", exc_info=True)
            return set()

def _get_shared_strings(z: zipfile.ZipFile) -> List[str]:
    """Extracts shared strings from the Excel file's sharedStrings.xml."""
    try:
        shared_strings_xml = z.read("xl/sharedStrings.xml")
        root = ET.fromstring(shared_strings_xml)
        ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        strings = []
        for sst_item in root.findall(".//ns:si", ns):
            t_element = sst_item.find("ns:t", ns)
            if t_element is not None:
                strings.append(t_element.text if t_element.text is not None else "")
            else:
                full_text = []
                for r_element in sst_item.findall("ns:r", ns):
                    r_t_element = r_element.find("ns:t", ns)
                    if r_t_element is not None:
                        full_text.append(r_t_element.text if r_t_element.text is not None else "")
                strings.append("".join(full_text))
        return strings
    except KeyError:
        logging.warning("sharedStrings.xml not found in the Excel file. No shared strings to load.")
        return []
    except Exception as e:
        logging.error(f"Error loading shared strings: {e}", exc_info=True)
        return []

def _extract_cell_values_from_xml(z: zipfile.ZipFile, shared_strings: List[str], sheet_name: str) -> Dict[Tuple[int, int], Any]:
    """Extracts cell values for a specific sheet by parsing its XML."""
    cell_values_map: Dict[Tuple[int, int], Any] = {}

    excel_error_strings = {"#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!"}
    try:
        xml_path = _find_sheet_xml_path(z, sheet_name)
        tree = ET.fromstring(z.read(xml_path))
        ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

        for row_elem in tree.findall(".//ns:row", ns):
            row_num_str = row_elem.attrib.get("r")
            if not row_num_str:
                continue
            try:
                row_num = int(row_num_str)
            except ValueError:
                logging.warning(f"Invalid row number '{row_num_str}' in sheet '{sheet_name}'. Skipping row.")
                continue

            for cell_elem in row_elem.findall(".//ns:c", ns):
                cell_ref = cell_elem.attrib.get("r")
                if not cell_ref:
                    continue
                try:
                    _, col_num = coordinate_to_tuple(cell_ref)
                except ValueError:
                    logging.warning(f"Invalid cell reference '{cell_ref}' in sheet '{sheet_name}'. Skipping cell.")
                    continue

                cell_type = cell_elem.attrib.get("t", "n")
                value_elem = cell_elem.find("ns:v", ns)
                value = None
                
                if value_elem is not None and value_elem.text is not None:
                    if value_elem.text in excel_error_strings:
                        value = 0
                    elif cell_type == "s":
                        try:
                            s_idx = int(value_elem.text)
                            if 0 <= s_idx < len(shared_strings):
                                value = shared_strings[s_idx]
                            else:
                                logging.warning(f"Shared string index {s_idx} out of bounds for cell {cell_ref} in sheet {sheet_name}.")
                        except ValueError:
                            logging.warning(f"Invalid shared string index '{value_elem.text}' for cell {cell_ref}.")
                    elif cell_type == "b":
                        value = bool(int(value_elem.text))
                    elif cell_type == "n":
                        try:
                            value = float(value_elem.text)
                        except ValueError:
                            value = value_elem.text
                    elif cell_type == "str":
                        value = value_elem.text
                    elif cell_type == "inlineStr":
                        is_elem = cell_elem.find("ns:is", ns)
                        if is_elem is not None:
                            t_elem = is_elem.find("ns:t", ns)
                            if t_elem is not None:
                                value = t_elem.text
                    else:
                        value = value_elem.text

                if value is not None:
                    cell_values_map[(row_num, col_num)] = value

        return cell_values_map
    except KeyError as e:
        logging.error(f"Could not find XML path for sheet '{sheet_name}' or related entry: {e}")
        return {}
    except Exception as e:
        logging.error(f"Error extracting cell values for sheet '{sheet_name}': {e}", exc_info=True)
        return {}

def preprocess_excel_xml(source_file: str, target_file: str):
    """
    Copies specific sheets from a source Excel file to a target Excel file using XML parsing.
    Dynamically determines the required sheets based on fiscal calendar and naming patterns,
    and preserves hidden row settings.
    """
    logging.info(f"Starting Excel preprocessing: Source file = {source_file}, Target file = {target_file}")

    current_date = datetime.now()
    fiscal_year, fiscal_quarter_str, fiscal_month_overall_str, fiscal_month_in_quarter_str = get_fiscal_quarter_and_month(current_date.date())
    logging.info(f"Current fiscal period for preprocessing: {fiscal_quarter_str} {fiscal_month_overall_str} (Month in Q: {fiscal_month_in_quarter_str})")

    with zipfile.ZipFile(source_file, 'r') as z:
        ns_wb = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        wb_xml = ET.fromstring(z.read("xl/workbook.xml"))
        sheet_elements = wb_xml.findall(".//ns:sheet", ns_wb)
        all_sheet_names = [sheet.attrib["name"] for sheet in sheet_elements]

        required_sheets = find_dynamic_sheets(all_sheet_names, fiscal_quarter_str, fiscal_month_overall_str, fiscal_month_in_quarter_str)
        logging.info(f"Sheets identified for preprocessing: {required_sheets}")

        if len(required_sheets) < 4:
            logging.warning("Not all expected sheets were found for preprocessing. Some may be missing.")

        tgt_wb = Workbook()
        if "Sheet" in tgt_wb.sheetnames:
            del tgt_wb["Sheet"] # Remove default sheet

        shared_strings = _get_shared_strings(z)

        for sheet_name in required_sheets:
            logging.info(f"Processing sheet for preprocessing: {sheet_name}")
            try:
                hidden_rows = _get_hidden_rows_from_xml(source_file, sheet_name)
                cell_values_map = _extract_cell_values_from_xml(z, shared_strings, sheet_name)

                if not cell_values_map and not hidden_rows:
                    logging.warning(f"Sheet '{sheet_name}' appears empty or could not be processed during preprocessing. Skipping.")
                    continue

                tgt_sheet = tgt_wb.create_sheet(title=sheet_name)
                logging.info(f"Sheet '{sheet_name}' created in target workbook for preprocessing.")

                max_row, max_col = 0, 0
                for (row_num, col_num), cell_value in cell_values_map.items():
                    tgt_sheet.cell(row=row_num, column=col_num, value=cell_value)
                    max_row, max_col = max(max_row, row_num), max(max_col, col_num)

                for row_idx in hidden_rows:
                    if row_idx <= max_row:
                        tgt_sheet.row_dimensions[row_idx].hidden = True
                logging.info(f"Applied hidden row settings for sheet: {sheet_name}")

            except Exception as e:
                logging.error(f"Error processing sheet '{sheet_name}' during preprocessing: {e}", exc_info=True)
        try:
            tgt_wb.save(target_file)
            logging.info(f"Preprocessed Excel saved to {target_file}")
        except Exception as e:
            logging.critical(f"Error saving target workbook '{target_file}': {e}", exc_info=True)