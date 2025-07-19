import time
import glob
import logging
from datetime import datetime
from openpyxl import load_workbook

import config
from utils import get_dynamic_filename_components, get_fiscal_quarter_and_month, find_dynamic_sheets
from excel_processor import preprocess_excel_xml
from ppt_updater import load_tables_from_excel, update_ppt_labels

def main():
    start_time = time.time()
   
    # --- Step 0: Determine source Excel file dynamically ---
    future_prefix, today_date_str = get_dynamic_filename_components()
    logging.info(f"Calculated future file prefix: '{future_prefix}'. Today's date string (for '{{date}}'): '{today_date_str}'.")

    source_excel = None

    # Primary Search Attempt: Look for files with both future prefix and today's date
    primary_search_pattern = f"{future_prefix}*.xlsm"
    logging.info(f"Attempting primary search using pattern: '{primary_search_pattern}'.")
    found_xlsm_files = [f for f in glob.glob(primary_search_pattern) if f != config.TARGET_EXCEL_FILENAME]

    # Fallback Search Attempt: If no file found with today's date, try only with future prefix
    if not found_xlsm_files:
        logging.warning(f"No .xlsm file found with primary pattern '{primary_search_pattern}'. Trying fallback pattern: '{future_prefix}*.xlsm'.")
        fallback_search_pattern = f"{future_prefix}*.xlsm"
        found_xlsm_files = [f for f in glob.glob(fallback_search_pattern) if f != config.TARGET_EXCEL_FILENAME]

    # Handle Search Results
    if not found_xlsm_files:
        logging.critical(f"No .xlsm file found matching any criteria in the current directory (excluding '{config.TARGET_EXCEL_FILENAME}'). Exiting.")
        return
    elif len(found_xlsm_files) > 1:
        logging.warning(f"Multiple .xlsm files found: {found_xlsm_files}. Using the first one: '{found_xlsm_files[0]}'.")
        source_excel = found_xlsm_files[0]
    else:
        source_excel = found_xlsm_files[0]
   
    logging.info(f"Dynamically determined source_excel: '{source_excel}'.")
   
    # --- Step 1: Preprocess the Excel file to create the cleaned/preprocessed Excel ---
    preprocess_excel_xml(source_excel, config.TARGET_EXCEL_FILENAME)

    # --- Step 2: Load the preprocessed Excel file and prepare for PowerPoint update ---
    # Load workbook with data_only=True to get cell values (not formulas)
    wb_data_only = load_workbook(config.TARGET_EXCEL_FILENAME, data_only=True, read_only=True)
   
    sheet_names = wb_data_only.sheetnames
    logging.info(f"Sheets in preprocessed Excel: {sheet_names}")

    current_date = datetime.now()
    fiscal_year, fiscal_quarter_str, fiscal_month_overall_str, fiscal_month_in_quarter_str = get_fiscal_quarter_and_month(current_date.date())
    logging.info(f"Current fiscal period for PPT update: {fiscal_quarter_str} {fiscal_month_overall_str} (Month in Q: {fiscal_month_in_quarter_str})")

    # --- Dynamic PowerPoint filename generation ---
    # Example: M3 Q4FY25 P&L Review_Cisco Highly Confidential _WD-1 DRAFT
    fiscal_year_short = str(fiscal_year)[2:] # Get last two digits of the fiscal year (e.g., '25' from '2025')
    dynamic_filename_part = f"{fiscal_month_in_quarter_str} {fiscal_quarter_str}FY{fiscal_year_short}"
    
    # Construct the full new PowerPoint filename
    new_ppt_filename = f"{dynamic_filename_part} P&L Review_Cisco Highly Confidential _WD-1 DRAFT.pptx"
    
    # Update the final output path using the Path object from config
    dynamic_final_output_ppt_filename = config.OUTPUT_DIRECTORY / new_ppt_filename
    logging.info(f"Dynamic final PowerPoint output filename set to: '{dynamic_final_output_ppt_filename}'.")

    # Dynamically determine the required sheet names for PPT update
    identified_sheets = find_dynamic_sheets(sheet_names, fiscal_quarter_str, fiscal_month_overall_str, fiscal_month_in_quarter_str)
    logging.info(f"Sheets identified for PPT update: {identified_sheets}")

    # Define the desired order of sheets for table processing
    desired_order = ["Exec View", "Comparisons", "Commit", "Margins Scenarios"]
    ordered_sheets = []

    # Map dynamically identified sheets to their corresponding positions in the desired order
    for desired_name in desired_order:
        matched_sheet = None
        for sheet in identified_sheets:
            if desired_name in sheet:
                matched_sheet = sheet
                break
        if matched_sheet:
            ordered_sheets.append(matched_sheet)
        else:
            logging.warning(f"Sheet matching '{desired_name}' not found for PPT update. It will be skipped.")

    logging.info(f"Ordered sheets for PPT update: {ordered_sheets}")

    if len(ordered_sheets) < 4:
        missing_sheets = set(desired_order) - set(s.split(' ', 1)[-1] for s in ordered_sheets) # Adjust to compare just the 'Exec View' part
        logging.critical(f"Missing critical sheets for PPT update: {', '.join(missing_sheets)}. Ensure the source file is correct and named as expected.")

    # Define the base template for table regions
    base_regions = {
        "Exec View": [
            ("C3", "E13"), ("F3", "F13"),
            ("C18", "E24"), ("F18", "F24"), ("H18", "H18"), ("H20", "H22"),
            ("C29", "E36"), ("F29", "F36"), ("H29", "H29"), ("H31", "H33"),
            ("K3", "K3"), ("K4", "N13"), ("O3", "O3"), ("R4", "R13"),
            ("S3", "S3"), ("V4", "V13"), ("W3", "W3"), ("W4", "W13"),
        ],
        "Comparisons": [
            ("K3", "K3"), ("K4", "N13"), ("S3", "S13"), ("T3", "T3"),
            ("W4", "W13"), ("X3", "X13"), ("AC3", "AC13"), ("AD3", "AD3"),
            ("AG4", "AG13"), ("AH3", "AH13"), ("AM3", "AM13"), ("AN3", "AN3"), ("AQ4", "AQ13"),
            ("AR3", "AR13"),
        ],
        "Commit": [
            ("C3", "C3"), ("C4", "F13"), ("G3", "G3"), ("J4", "J13"),
            ("K3", "K3"), ("N4", "N13"), ("O3", "O3"), ("R4", "R13"),
            ("S3", "S3"), ("V4", "V13"),
        ],
        "Margins Scenarios": [
            ("B15", "B15"), ("B16", "G19"), ("B20", "G20"), ("B25", "B25"),
            ("B26", "G29"), ("B30", "G30"), ("B32", "B32"), ("B33", "G36"),
            ("B37", "G37"), ("B39", "B39"), ("B40", "G43"), ("B44", "G44"),
            ("B46", "B46"), ("B47", "G50"), ("B51", "G51"), ("I39", "I39"), 
            ("I40", "N43"), ("I44", "N44"),
        ],
    }

    # Dynamically map sheet names to their corresponding regions
    table_regions = {}
    for sheet_name in ordered_sheets:
        if "Exec View" in sheet_name:
            table_regions[sheet_name] = base_regions["Exec View"]
        elif "Comparisons" in sheet_name:
            table_regions[sheet_name] = base_regions["Comparisons"]
        elif "Commit" in sheet_name:
            table_regions[sheet_name] = base_regions["Commit"]
        elif "Margins Scenarios" in sheet_name:
            table_regions[sheet_name] = base_regions["Margins Scenarios"]

    logging.info("Loading tables from preprocessed Excel for PowerPoint update...")
    tables = load_tables_from_excel(str(config.TARGET_EXCEL_FILENAME), table_regions.keys(), table_regions, wb_data_only)
    # In main.py, after tables = load_tables_from_excel(...)
    logging.info(f"Loaded tables keys: {tables.keys()}")
    # Assuming your Commit sheet is named 'Q4 Commit' for example
    commit_sheet_name = next((s for s in ordered_sheets if "Commit" in s), None)
    if commit_sheet_name:
        commit_tables = tables.get(commit_sheet_name)
        if commit_tables and len(commit_tables) > 5: # Check if the 6th table exists
            ah_table_df = commit_tables[5] # Index 5 for the 6th table
            logging.info(f"DataFrame for ah (Commit sheet, region N4:N13) shape: {ah_table_df.shape}")
            logging.info(f"Value for ah9 (N12): {ah_table_df.iat[8, 0]}")
        else:
            logging.warning(f"Commit sheet tables not found or not enough tables for ah.")
    logging.info("Tables loaded successfully.")

    # --- Step 3: Update the PowerPoint presentation with the loaded data ---
    logging.info("Updating PowerPoint presentation...")
    # Pass the dynamically generated filename to update_ppt_labels
    update_ppt_labels(str(config.PPT_TEMPLATE_FILENAME), str(dynamic_final_output_ppt_filename), tables)

    logging.info(f"PowerPoint updated and saved to {dynamic_final_output_ppt_filename}")
    end_time = time.time()
    elapsed_time = end_time - start_time
    logging.info(f"Script completed in {elapsed_time:.2f} seconds.")

if __name__ == "__main__":
    main()